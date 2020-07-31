from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl import load_workbook
from termcolor import colored


def add_style_survie(survie_dataframe, PATH):

    wb = load_workbook(PATH)
    ws = wb['Survie']

    nb_rows, nb_columns = survie_dataframe.shape

    medium = Side(border_style='medium', color='FF000000')
    medium_borders = Border(top=medium, left=medium, right=medium, bottom=medium)
    font_Arial9 = Font(size=9, name='Arial')
    alignment_center = Alignment(horizontal='center', vertical='center')

    ## COLUMN WIDTH ##
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 3
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 6
    ws.column_dimensions['E'].width = 45
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 20

    ## HEADER STYLE ##
    merging_cells = [ 'C2:C3', 'D2:D3', 'E2:E3', 'F2:F3', 'G2:G3']
    border_cells = [ 'C2', 'C3', 'D2', 'D3', 'E2', 'E3', 'F2', 'F3', 'G2', 'G3']
    merging_names = ['Campagne', '#', 'Station de mesure', 'Code Agence', 'Survie biotest chimie']

    header_stations_font = Font(size=10, bold=True, name='Arial', color='FFFFFF')
    header_stations_fill = PatternFill(fill_type='solid', start_color='808080', end_color='808080')

    for i in range(len(merging_cells)):
        cells = merging_cells[i]
        top_left_cell = ws[cells[:2]]
        name = merging_names[i]

        ws.merge_cells(cells)
        top_left_cell.value = name
        top_left_cell.font = header_stations_font
        top_left_cell.fill = header_stations_fill
        top_left_cell.alignment = alignment_center

    for cell in border_cells:
        ws[cell].border = medium_borders

    ## BODY STYLE ##
    body_rows = [str(r) for r in range(4, nb_rows + 4)]
    body_columns = [ 'C', 'D', 'E', 'F', 'G']
    body_cells = []
    for row in body_rows:
        for column in body_columns:
            body_cells.append(column + row)

    for cell_str in body_cells:
        cell = ws[cell_str]
        value = cell.value

        if value == '' or value is None:
            cell.value = 'ND'
        if cell_str[0] == 'H':
            cell.font = Font(size=9, name='Arial', italic=True)
        else:
            cell.font = font_Arial9
        cell.alignment = alignment_center
        cell.border = medium_borders
    ws.freeze_panes = ws["A4"]
    wb.save(PATH)
    wb.close()

