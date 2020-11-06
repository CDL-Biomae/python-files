from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter
import pandas as pd
from calcul import chemistry
from tools import QueryScript
from termcolor import colored
import env

def add_style_bbac_21j(bbac_dataframe, PATH, dict_t0, old_color):

    wb = load_workbook(PATH)
    ws = wb['BBAC_21j']
    ws2 = wb['BBAC2_21j']
    
    nb_rows, nb_columns = bbac_dataframe.shape
    header_row = '4'
    header_columns = [get_column_letter(col_idx) for col_idx in list(range(2, nb_columns + 2))]
    borders = Border(left=Side(border_style='thin', color='FFFFFF'),
                     right=Side(border_style='thin', color='FFFFFF'),
                     top=Side(border_style='thin', color='FFFFFF'),
                     bottom=Side(border_style='thin', color='FFFFFF'))
     
    elements_metal = QueryScript(f" SELECT sandre, parameter, 21j_threshold, 21j_graduate_25, 21j_graduate_50, 21j_graduate_75   FROM {env.DATABASE_TREATED}.r3 WHERE version=  {env.CHOSEN_VERSION()} and familly='Métaux' AND 21j_threshold IS NOT NULL").execute()
    elements_organic = QueryScript(f" SELECT sandre, parameter, 21j_threshold, 21j_graduate_25, 21j_graduate_50, 21j_graduate_75   FROM {env.DATABASE_TREATED}.r3 WHERE version=  {env.CHOSEN_VERSION()} and familly!='Métaux' AND 21j_threshold IS NOT NULL").execute()
    
    for letter in [get_column_letter(col_idx) for col_idx in range(1, nb_columns+5)]:
        for number in range(1, nb_rows+21):
            ws[letter+str(number)].border = borders
            ws2[letter+str(number)].border = borders
    
    ## UNIT 
    [unit_metal, sandre_metal] = chemistry.get_unit([int(float(element[0])) for element in elements_metal]) 
    parameter_metal = [element[1] for element in elements_metal]
    threshold_metal = [element[2:] for element in elements_metal]
    [unit_organic, sandre_organic] = chemistry.get_unit([int(float(element[0])) for element in elements_organic]) 
    parameter_organic = [element[1] for element in elements_organic]
    threshold_organic = [element[2:] for element in elements_organic]
    
    for letter in header_columns[5:]:
        index =None
        if ws[letter + '4'].value and int(ws[letter + '4'].value) in sandre_metal:
            index = sandre_metal.index(int(ws[letter + '4'].value))
            sandre_checked = sandre_metal
            parameter_checked = parameter_metal
            unit_checked = unit_metal
        elif ws[letter + '4'].value and int(ws[letter + '4'].value) in sandre_organic:
            index = sandre_organic.index(int(ws[letter + '4'].value))
            sandre_checked = sandre_organic
            parameter_checked = parameter_organic
            unit_checked = unit_organic
        if index!=None:
            ws[letter + '2'].value = unit_checked[index]
            ws2[letter + '2'].value = unit_checked[index]
            ws[letter + '3'].value = sandre_checked[index]
            ws2[letter + '3'].value = sandre_checked[index]
            ws[letter + '4'].value = parameter_checked[index]
            ws2[letter + '4'].value = parameter_checked[index]


    
            
    
    ## HEADER STYLE ##
    
    ws['B2'].value = 'Campagne'
    ws2['B2'].value = 'Campagne'
    ws['C2'].value = '#'
    ws2['C2'].value = '#'
    ws['D2'].value = 'Station de mesure'
    ws2['D2'].value = 'Station de mesure'
    ws['E2'].value = 'Code agence'
    ws2['E2'].value = 'Code agence'
    
    header_cells = [c+header_row for c in header_columns]
    header_font = Font(size=8, bold=True, name='Arial')
    header_alignment_rotate = Alignment(horizontal='center', vertical='bottom', text_rotation=90)
    header_alignment_no_rotate = Alignment(horizontal='center', vertical='bottom')
    borders = Border(left=Side(border_style='thin', color='FF000000'),
                     right=Side(border_style='thin', color='FF000000'),
                     top=Side(border_style='thin', color='FF000000'),
                     bottom=Side(border_style='thin', color='FF000000'))
    
    ws.column_dimensions['B'].width=3
    ws2.column_dimensions['B'].width=3
    ws.column_dimensions['C'].width=3
    ws2.column_dimensions['C'].width=3
    ws.column_dimensions['D'].width=30
    ws2.column_dimensions['D'].width=30
    ws.column_dimensions['E'].width=8
    ws2.column_dimensions['E'].width=8
    for letter in header_columns[:4]:
        for number in range(2,5):
            ws[letter+str(number)].border = borders
            ws2[letter+str(number)].border = borders
            ws[letter+str(number)].font = header_font
            ws2[letter+str(number)].font = header_font
    ws['B2'].alignment = header_alignment_rotate
    ws['C2'].alignment = header_alignment_no_rotate
    ws['D2'].alignment = header_alignment_no_rotate
    ws['E2'].alignment = header_alignment_rotate
    ws2['B2'].alignment = header_alignment_rotate
    ws2['C2'].alignment = header_alignment_no_rotate
    ws2['D2'].alignment = header_alignment_no_rotate
    ws2['E2'].alignment = header_alignment_rotate
    
    
    
    header_font = Font(size=8, name='Arial')
    
    for letter in header_columns[4:]:
        if (ws[letter+'5'].value !=None and ws[letter+'5'].value !='') :
            
            ws.column_dimensions[letter].width=6
            ws[letter+'4'].alignment = header_alignment_rotate
            ws[letter+'4'].font = header_font
            ws[letter+'3'].alignment = header_alignment_no_rotate
            ws[letter+'3'].font = header_font
            ws[letter+'2'].alignment = header_alignment_no_rotate
            ws[letter+'2'].font = header_font
            ws2.column_dimensions[letter].width=4
            ws2[letter+'4'].alignment = header_alignment_rotate
            ws2[letter+'4'].font = header_font
            ws2[letter+'3'].alignment = header_alignment_no_rotate
            ws2[letter+'3'].font = header_font
            ws2[letter+'2'].alignment = header_alignment_no_rotate
            ws2[letter+'2'].font = header_font
        else :
            ws.column_dimensions[letter].width=2
            ws2.column_dimensions[letter].width=2
            
 ## ADD T0
    
    t0_mp = []
    for mp in dict_t0:
        if not dict_t0[mp]['code_t0_id'] in t0_mp and dict_t0[mp]['code_t0_id']:
            t0_mp.append(dict_t0[mp]['code_t0_id'])
    if len(t0_mp) > 1 or len(t0_mp) == 0:
        query_tuple_t0 = tuple(t0_mp)
    else:
        query_tuple_t0 = f"({t0_mp[0]})"
    reference_dict = {}
    if len(t0_mp):
        reference_result = QueryScript(f"SELECT reference, id FROM {env.DATABASE_RAW}.Measurepoint WHERE id IN {query_tuple_t0}").execute()
    else:
        reference_result = []
    for reference in reference_result:
        reference_dict.update({reference[1]:reference[0]})
    t0_result=[]
    if len(t0_mp):
        t0_result = QueryScript(f"SELECT sandre, prefix, value, Pack.measurepoint_id, Measurepoint.reference FROM {env.DATABASE_RAW}.Analysis JOIN {env.DATABASE_RAW}.Pack ON Pack.id= Analysis.pack_id JOIN {env.DATABASE_RAW}.Measurepoint ON Pack.measurepoint_id=Measurepoint.id WHERE Pack.measurepoint_id IN {query_tuple_t0};").execute()
    dict_t0_result= {}
    for element in t0_result:
        if not element[3] in dict_t0_result:
            dict_t0_result.update({element[3]: {element[0]:element[1] + str(element[2]) if element[1] else str(element[2]), 'reference': element[4]}})
        else :
            dict_t0_result[element[3]][element[0]] = element[1] + str(element[2]) if element[1] else str(element[2])
    
    t0_font = Font(size=6, name='Arial')
    t0_border = Border(left=Side(border_style='thin', color='000000'),
                     right=Side(border_style='thin', color='000000'),
                     top=Side(border_style='thin', color='000000'),
                     bottom=Side(border_style='thin', color='000000'))
    body_alignment = Alignment(horizontal='center', vertical='center')
    
    for index, t0 in enumerate(t0_mp):
        ws['B'+str(nb_rows+5+index)].font = t0_font
        ws['B'+str(nb_rows+5+index)].border = t0_border
        ws['C'+str(nb_rows+5+index)].font = t0_font
        ws['C'+str(nb_rows+5+index)].border = t0_border
        ws['D'+str(nb_rows+5+index)].font = t0_font
        ws['D'+str(nb_rows+5+index)].border = t0_border
        if t0 in dict_t0_result :
            ws['D'+str(nb_rows+5+index)].value = dict_t0_result[t0]['reference']
            ws2['D'+str(nb_rows+5+index)].value = dict_t0_result[t0]['reference']
        else :
            ws['D'+str(nb_rows+5+index)].value = reference_dict[t0]
            ws2['D'+str(nb_rows+5+index)].value = reference_dict[t0]
        ws['E'+str(nb_rows+5+index)].font = t0_font
        ws['E'+str(nb_rows+5+index)].border = t0_border
        ws2['B'+str(nb_rows+5+index)].font = t0_font
        ws2['B'+str(nb_rows+5+index)].border = t0_border
        ws2['C'+str(nb_rows+5+index)].font = t0_font
        ws2['C'+str(nb_rows+5+index)].border = t0_border
        ws2['D'+str(nb_rows+5+index)].font = t0_font
        ws2['D'+str(nb_rows+5+index)].border = t0_border
        ws2['E'+str(nb_rows+5+index)].font = t0_font
        ws2['E'+str(nb_rows+5+index)].border = t0_border
        for letter in header_columns[5:]:
            sandre = ws[letter +'3'].value
            if t0 in dict_t0_result and  str(sandre) in dict_t0_result[t0]:
                ws[letter+str(nb_rows+5+index)].value = dict_t0_result[t0][str(sandre)]
                ws[letter+str(nb_rows+5+index)].font = t0_font
                ws[letter+str(nb_rows+5+index)].border = t0_border
                ws[letter+str(nb_rows+5+index)].alignment = body_alignment
                ws2[letter+str(nb_rows+5+index)].value = dict_t0_result[t0][str(sandre)]
                ws2[letter+str(nb_rows+5+index)].font = t0_font
                ws2[letter+str(nb_rows+5+index)].border = t0_border
                ws2[letter+str(nb_rows+5+index)].alignment = body_alignment
                
            elif sandre != None :
                ws[letter+str(nb_rows+5+index)].value = 'ND'
                ws[letter+str(nb_rows+5+index)].font = t0_font
                ws[letter+str(nb_rows+5+index)].border = t0_border
                ws[letter+str(nb_rows+5+index)].alignment = body_alignment
                ws2[letter+str(nb_rows+5+index)].font = t0_font
                ws2[letter+str(nb_rows+5+index)].border = t0_border
                ws2[letter+str(nb_rows+5+index)].alignment = body_alignment
        

    ## BODY STYLE ##
    body_rows = [str(r) for r in list(range(5, nb_rows + 5 + len(t0_mp)))]
    body_columns = header_columns[5:]

    body_font = Font(size=6, name='Arial')
    body_font_white = Font(size=6, name='Arial', color='FFFFFF')
    body_fill_ok = PatternFill(fill_type='solid', start_color='DBB7FF' if not old_color else '027ee3', end_color='DBB7FF' if not old_color else '027ee3')
    body_fill_nd = PatternFill(fill_type='solid', start_color='A6A6A6' , end_color='A6A6A6' )
    body_fill_not_ok = PatternFill(fill_type='solid', start_color='B565F7' if not old_color else '69a64b', end_color='B565F7' if not old_color else '69a64b')
    body_fill_not_ok_25 = PatternFill(fill_type='solid', start_color='8909FF' if not old_color else 'd1c452', end_color='8909FF' if not old_color else 'd1c452')
    body_fill_not_ok_50 = PatternFill(fill_type='solid', start_color='6600CC' if not old_color else 'cc7931', end_color='6600CC' if not old_color else 'cc7931')
    body_fill_not_ok_75 = PatternFill(fill_type='solid', start_color='47008E' if not old_color else 'ab2222', end_color='47008E' if not old_color else 'ab2222')
    body_alignment = Alignment(horizontal='center', vertical='center')

    for column in header_columns:
        for row in body_rows:
            if column in ['B','C','D','E']:
                ws[column+row].border = borders
                ws[column+row].font = body_font
                ws2[column+row].border = borders
                ws2[column+row].font = body_font


    for column in body_columns:
        sandre_checked = ws[column+'3'].value
        threshold_21j=''
        if sandre_checked!='' and sandre_checked!=None:
            try :
                index = sandre_metal.index(sandre_checked)
                threshold_21j = threshold_metal[index][0]
                threshold_21j_25 = threshold_metal[index][1]
                threshold_21j_50 = threshold_metal[index][2]
                threshold_21j_75 = threshold_metal[index][3]
            except :
                index = sandre_organic.index(sandre_checked)
                threshold_21j = threshold_organic[index][0]
                threshold_21j_25 = threshold_organic[index][1]
                threshold_21j_50 = threshold_organic[index][2]
                threshold_21j_75 = threshold_organic[index][3]
        for row in body_rows:
            cell = ws[column+row]
            cell2 = ws2[column+row]
            value = cell.value
            cell.font = body_font
            cell2.font = body_font
            if value!=None :
                cell.alignment = body_alignment
                cell.border = borders
                cell2.alignment = body_alignment
                cell2.border = borders
                if (value!="ND" and value!='0.0' and threshold_21j!='') and ((value!='' and value[0]=='<') or float(value)<threshold_21j):
                    cell.fill = body_fill_ok
                    cell2.fill = body_fill_ok
                    cell2.value = 0
                elif (value!="ND" and value!='0.0' and threshold_21j!='' and float(value)>=threshold_21j):
                    if threshold_21j_25==None:
                        cell.fill = body_fill_not_ok_75
                        cell.font = body_font_white
                        cell2.fill = body_fill_not_ok_75
                        cell2.value = 4
                        cell2.font = body_font_white
                    elif float(value)>=threshold_21j_25 :
                        if float(value)>=threshold_21j_50:
                            if float(value)>=threshold_21j_75:
                                cell.fill = body_fill_not_ok_75
                                cell.font = body_font_white
                                cell2.value = 4
                                cell2.fill = body_fill_not_ok_75
                                cell2.font = body_font_white
                            else :    
                                cell.fill = body_fill_not_ok_50 
                                cell.font = body_font_white
                                cell2.fill = body_fill_not_ok_50 
                                cell2.value = 3
                                cell2.font = body_font_white
                        else:
                            cell.fill = body_fill_not_ok_25 
                            cell2.fill = body_fill_not_ok_25 
                            cell2.value = 2
                    else :
                        cell.fill = body_fill_not_ok
                        cell2.fill = body_fill_not_ok
                        cell2.value = 1
                else :
                    cell.fill = body_fill_nd
                    cell2.fill = body_fill_nd
                    cell2.value = 'nd'
    for index,mp in enumerate(dict_t0):
        try :
            if ws[header_columns[-1] + str(index+5)].value and ws[header_columns[-1] + str(index+5)].value in dict_t0:
                index_t0_associated = t0_mp.index(dict_t0[ws[header_columns[-1] + str(index+5)].value]['code_t0_id'])
                for column in header_columns[5:]:
                    t0_ok = True if ws[column + str(5+nb_rows+index_t0_associated)].fill == body_fill_ok else False 
                    if not t0_ok and (ws[column + str(5+nb_rows+index_t0_associated)].value!= None and ws[column + str(5+nb_rows+index_t0_associated)].value!= ''):
                        ws[column + str(5+index)].fill = body_fill_nd
                        ws[column + str(5+index)].font = body_font
                        ws2[column + str(5+index)].fill = body_fill_nd
                        ws2[column + str(5+index)].font = body_font
                        ws2[column + str(5+index)].value = 'nd'
        except ValueError :
            None
    ws.delete_cols(len(header_columns)+1,1)
    ws2.delete_cols(len(header_columns)+1,1)
    
    for letter in header_columns[5:]:
        for number in range(5, nb_rows+21):
                ws[letter + str(number)].value = str(ws[letter + str(number)].value).replace(".", ",") if ws[letter + str(number)].value else ''
    
    ws.freeze_panes = ws["F8"]
    ws2.freeze_panes = ws2["F5"]

    wb.save(PATH)
    wb.close()
    wb = load_workbook(PATH)
    ws = wb['BBAC_21j']
    ws2 = wb['BBAC2_21j']
    ## Add threshold table
    ws.insert_rows(1)
    ws.insert_rows(1)
    ws.insert_rows(1)
    ws.insert_rows(1)

    ws["F1"].value = "Seuil"
    ws["F2"].value = "Seuil gradué 25%"
    ws["F3"].value = "Seuil gradué 50%"
    ws["F4"].value = "Seuil gradué 75%"
    
    ws.column_dimensions['F'].width=20
    for letter in header_columns[5:]:
        if ws[letter + "7"].value :
            for sandre, parameter, threshold, graduate_25, graduate_50, graduate_75 in elements_metal :
                if str(sandre)==str(ws[letter + "7"].value) :
                    ws[letter+"1"].value = threshold
                    ws[letter+"1"].border = borders
                    ws[letter+"2"].value = graduate_25
                    ws[letter+"2"].border = borders
                    ws[letter+"3"].value = graduate_50
                    ws[letter+"3"].border = borders
                    ws[letter+"4"].value = graduate_75
                    ws[letter+"4"].border = borders
            for sandre, parameter, threshold, graduate_25, graduate_50, graduate_75 in elements_organic :
                if str(sandre)==str(ws[letter + "7"].value) :
                    ws[letter+"1"].value = threshold
                    ws[letter+"1"].border = borders
                    ws[letter+"2"].value = graduate_25
                    ws[letter+"2"].border = borders
                    ws[letter+"3"].value = graduate_50
                    ws[letter+"3"].border = borders
                    ws[letter+"4"].value = graduate_75
                    ws[letter+"4"].border = borders


    ## Merge unit
    
    current_unit = ws['G6'].value 
    first_letter = 'G'           
    last_letter = 'G'           
    index = 6
    while index <len(header_columns): 
        while index <len(header_columns) and ws[header_columns[index] + '6'].value == current_unit :   
            last_letter = header_columns[index]
            index +=1
        ws.merge_cells(first_letter + '6:'+last_letter+'6')
        ws2.merge_cells(first_letter + '2:'+last_letter+'2')
        if index<len(header_columns):
            first_letter = last_letter = header_columns[index]
        current_unit = ws[first_letter +'6'].value
        index+=1

    ws.merge_cells('B6:B8')
    ws2.merge_cells('B2:B4')
    ws.merge_cells('C6:C8')
    ws2.merge_cells('C2:C4')
    ws.merge_cells('D6:D8')
    ws2.merge_cells('D2:D4')
    ws.merge_cells('E6:E8')
    ws2.merge_cells('E2:E4')

    wb.save(PATH)
    wb.close()


