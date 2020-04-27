from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter

def add_style_edi(edi_dataframe, PATH):

    wb = load_workbook(PATH)
    ws = wb['Export EDI']

    nb_rows, nb_columns = edi_dataframe.shape
    header_columns = [get_column_letter(col_idx) for col_idx in list(range(1, nb_columns + 2))]
    medium = Side(border_style='thin', color='FF000000')
    medium_borders = Border(top=medium, left=medium, right=medium, bottom=medium)
    font_head_1 = Font(size=10, name='Arial', bold=True)
    font_head = Font(size=20, name='Arial')
    alignment_center = Alignment(horizontal='center', vertical='center', wrapText=True)
    ws.row_dimensions[1].height=30
    ws.freeze_panes = ws["C4"]

    # Color 
    red = PatternFill(fill_type='solid', start_color='FF0000', end_color='FF0000')
    orange_light = PatternFill(fill_type='solid', start_color='FED8B1', end_color='FED8B1') 
    orange_dark = PatternFill(fill_type='solid', start_color='FF8C00', end_color='FF8C00') 
    orange = PatternFill(fill_type='solid', start_color='FFA500', end_color='FFA500') 
    blue = PatternFill(fill_type='solid', start_color='87CEFA', end_color='87CEFA') 
    blue_light = PatternFill(fill_type='solid', start_color='B0E0E6', end_color='B0E0E6') 
    yellow_light = PatternFill(fill_type='solid', start_color='FFFFE0', end_color='FFFFE0') 
    yellow = PatternFill(fill_type='solid', start_color='FFFF00', end_color='FFFF00') 
    green_light = PatternFill(fill_type='solid', start_color='90EE90', end_color='90EE90') 


    for letter in [get_column_letter(col_idx) for col_idx in range(1, nb_columns+1)]:
        for number in range(1, nb_rows+4):
            ws[letter+str(number)].border = medium_borders
            ws[letter+str(number)].alignment = alignment_center
    

    ## Head merge and precision
    for index, letter in enumerate(header_columns[:12]):
        ws[letter+"1"].font = font_head_1
        ws[letter+"1"].value = ws[letter+"3"].value
        if index < 8 :
            ws[letter+"1"].fill = orange_light
        elif index < 9 :
            ws[letter+"1"].fill = orange_dark
        else :
            ws[letter+"1"].fill = orange


        ws.merge_cells(letter+"1:"+letter+"3")
        
    ws["M1"].value = "Encagement"
    ws["M1"].font = font_head
    ws["M1"].fill = blue
    ws.merge_cells("M1:AL1")
    
    for letter in ["M","N","O","P"]:
        ws[letter+"2"].font = font_head_1
        ws[letter+"2"].value = ws[letter+"3"].value
        ws[letter+"2"].fill = blue
        ws.merge_cells(letter+"2:"+letter+"3")
    
    ws["Q2"].value = "Mesure sur site"
    ws["Q2"].font = font_head_1
    ws["Q2"].fill = blue
    ws.merge_cells("Q2:U2")

    ws["V2"].value = "Mesure Environnementale"
    ws["V2"].font = font_head_1
    ws["V2"].fill = blue
    ws.merge_cells("V2:AL2")
   
    ws["AM1"].value = "Prélèvement"
    ws["AM1"].font = font_head
    ws["AM1"].fill = blue
    ws.merge_cells("AM1:BL1")
    
    for letter in ["AM","AN","AO","AP"]:
        ws[letter+"2"].value = ws[letter+"3"].value
        ws[letter+"2"].font = font_head_1
        ws[letter+"2"].fill = blue
        ws.merge_cells(letter+"2:"+letter+"3")
    
    ws["AQ2"].value = "Mesure sur site"
    ws["AQ2"].font = font_head_1
    ws["AQ2"].fill = blue    
    ws.merge_cells("AQ2:AU2")

    ws["AV2"].value = "Mesure Environnementale"
    ws["AV2"].font = font_head_1
    ws["AV2"].fill = blue    
    ws.merge_cells("AV2:BL2")
    
    for letter in ["BM","BN"] :
        ws[letter+"1"].font = font_head
        ws[letter+"1"].value = ws[letter+"3"].value
        ws[letter+"1"].fill = red   
        ws.merge_cells(letter+"1:"+letter+"3")
    for letter in ["BO","BP","BQ","BR","BS","BT"] :
        ws[letter+"2"].font = font_head_1
        ws[letter+"2"].value = ws[letter+"3"].value
        ws[letter+"2"].fill = blue   
        ws.merge_cells(letter+"2:"+letter+"3")

    ws["BO1"].value = "Echantillonage"
    ws["BO1"].font = font_head
    ws["BO1"].fill = blue
    ws.merge_cells("BO1:BR1")
    ws["BS1"].value = "Après Lyophilisation"
    ws["BS1"].font = font_head
    ws["BS1"].fill = blue
    ws.merge_cells("BS1:BT1")

    # column width
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 32
    ws.column_dimensions["F"].width = 32
    ws.column_dimensions["G"].width = 85
    ws.column_dimensions["I"].width = 20
    ws.column_dimensions["J"].width = 15
    ws.column_dimensions["K"].width = 15
    ws.column_dimensions["L"].width = 15

    for letter in header_columns[12:64]:
        ws.column_dimensions[letter].width = 15
        ws[letter + "3"].fill = blue
    ws.column_dimensions["BM"].width = 20
    for letter in header_columns[65:72]:
        ws.column_dimensions[letter].width = 30
        ws[letter + "3"].fill = blue
    # for letter in ["W","X","Y","Z","AA","AL","AM","AN","AO","AP","AQ"]:
    #     ws.column_dimensions[letter].width = 20
    #     if letter != "AQ":
    #         ws[letter + "3"].fill = blue
    ws.column_dimensions["BN"].width = 85

    for letter in ["H","N","AN"]:
        ws.column_dimensions[letter].width = 0.01
    # for letter in ["AS","AT","AU","AV","AW","AX"]:
    #     ws.column_dimensions[letter].width = 30

    not_validated_fill = PatternFill(fill_type='solid', start_color='FF0000', end_color='FF0000')
    not_validated_font = Font(size=10, name='Arial', bold=True, color="FFFFFF")

    for row in range(4,nb_rows+4):
        for index, letter in enumerate(header_columns):
            if letter == "A":
                ws[letter + str(row)].number_format="MMM-YY"
            if letter in ["J","K","L"]:
                ws[letter + str(row)].fill = yellow
            if letter in ["D","M","AM"]:
                ws[letter + str(row)].fill = red
            if letter in ["I","O","P","AO","AP","BM","BN","BO","BP","BQ","BR","BS","BT"]:
                ws[letter + str(row)].fill = blue_light
            if letter in ["Q","R","S","T","U","AQ","AR","AS","AT","AU"]:
                ws[letter + str(row)].fill = green_light
            if index in range(21,38) or index in range(47,64):
                ws[letter + str(row)].fill = yellow_light
        if ws["BQ"+str(row)].value and ws["BR"+str(row)].value :
            if ws["BR"+str(row)].value - ws["BQ"+str(row)].value < 2500:
                ws["B"+str(row)].font= not_validated_font
                ws["B"+str(row)].fill= not_validated_fill
                ws["BR"+str(row)].font= not_validated_font
                ws["BR"+str(row)].fill= not_validated_fill
                ws["BQ"+str(row)].font= not_validated_font
                ws["BQ"+str(row)].fill= not_validated_fill
        if ws["BO"+str(row)].value and ws["BP"+str(row)].value :
            if ws["BP"+str(row)].value - ws["BO"+str(row)].value < 500:
                ws["B"+str(row)].font= not_validated_font
                ws["B"+str(row)].fill= not_validated_fill
                ws["BP"+str(row)].font= not_validated_font
                ws["BP"+str(row)].fill= not_validated_fill
                ws["BO"+str(row)].font= not_validated_font
                ws["BO"+str(row)].fill= not_validated_fill
        if ws["BM"+str(row)].value=="0%":
            ws["B"+str(row)].font= not_validated_font
            ws["B"+str(row)].fill= not_validated_fill
            ws["BM"+str(row)].font= not_validated_font
            ws["BM"+str(row)].fill= not_validated_fill
            
    




    wb.save(PATH)
    wb.close()

