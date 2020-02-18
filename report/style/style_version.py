from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter
from termcolor import colored

def add_style_version(version_dataframe, list_campaigns, filename):
    PATH = f"output\\{filename}"
    wb = load_workbook(PATH)
    ws = wb['Version']

    thin = Side(style='thin', color='000000')
    thin_borders = Border(top=thin, left=thin, right=thin, bottom=thin)
    alignment_center = Alignment(horizontal='center', vertical='center')

    ## COLUMN WIDTH ##
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 13
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 50

    ## NOM DU TABLEAU ##
    ws.merge_cells('B2:D2')
    cell = ws['B2']
    name = ''
    for c in list_campaigns:
        name += f"_{c}"

    cell.value = f'[TABLEAU ANNEXE]{name}'
    cell.font = Font(name='Calibri', size=14, bold=True, color='B20000')

    ## STYLE DU TABLEAU ##
    cells_nom_colonne = ['B4', 'C4', 'D4']
    for cell_str in cells_nom_colonne:
        cell = ws[cell_str]
        cell.font = Font(name='Calibri', size=11, bold=True)
        cell.border = thin_borders
        cell.alignment = alignment_center

    cells_value = ['B5', 'C5', 'D5']
    for cell_str in cells_value:
        cell = ws[cell_str]
        cell.font = Font(name='Calibri', size=11)
        cell.border = thin_borders
        cell.alignment = alignment_center

    wb.save(PATH)
    wb.close()

    print(colored('[+] La mise en page de l\'onglet \"Version\" est terminée', 'green'))
