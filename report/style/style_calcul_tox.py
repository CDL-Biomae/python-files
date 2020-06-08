from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter
from termcolor import colored


def add_style_calcul_tox(calcul_tox_dataframe, PATH):

    wb = load_workbook(PATH)
    ws = wb["Calcul tox"]

    nb_rows, nb_columns = calcul_tox_dataframe.shape
    header_columns = [
        get_column_letter(col_idx) for col_idx in list(range(2, nb_columns + 2))
    ]

    ## COLUMN WIDTH ##
    ws.column_dimensions["A"].width = 40
    for letter in header_columns:
        ws.column_dimensions[letter].width = 20

    thin = Side(style="thin", color="000000")
    borders = Border(top=thin, bottom=thin)
    without_borders = Border(left=thin, right=thin, top=thin, bottom=thin)

    gray = PatternFill(fill_type='solid', start_color='C0C0C0', end_color='C0C0C0')
    orange = PatternFill(fill_type='solid', start_color='FFC000', end_color='FFC000')
    yellow = PatternFill(fill_type='solid', start_color='FFFF00', end_color='FFFF00')
    green = PatternFill(fill_type='solid', start_color='8FBC8F', end_color='8FBC8F')
    blue = PatternFill(fill_type='solid', start_color='BDDAEF', end_color='BDDAEF')
    red = PatternFill(fill_type='solid', start_color='FF0000', end_color='FF0000')

    bold = Font(size=12, name='Calibri', bold=True)
    italic = Font(size=9, name='Calibri',italic=True)
    white = Font(size=11, name='Calibri', color="FFFFFF")

    for row in range(1,nb_rows) :
        row = str(row)
        if ws["A"+row].value in ["Survie 7 jours", "Alimentation", "Neurotoxicité - Activité AChE", "Reprotoxicité"]:
            ws["A"+row].fill = blue
            ws["A"+row].font = bold
            for letter in header_columns :
                ws[letter+row].fill = blue
                ws[letter+row].font = bold
        if not ws["A"+row].value :
            ws["A"+row].fill = blue
            ws["A"+row].font = bold
            for letter in header_columns :
                ws[letter+row].fill = blue
        if ws["A"+row].value in ["Moyenne de survie Alim 1 à 4 / Nb par réplictats","Moyenne de survie Alim X ; Nb par réplictats","Taille pixels - X x Contrôle étalon mm/pxl ","Etalon taille mm /  Etalon taille pixel","Somme des valeurs brutes pour 10 (témoin) / Nombre de disques (témoin)","Somme des valeurs brutes pour 20 (témoin) - Somme des valeurs pour réplicat X","Constante alim 1 x TEMPERATURE MOYENNE + Constante alim 2 + Constante alim 3 x (TAILLE MALE - Constante 4)","(Valeur d'alimentation attendue - mm² consommés/jour/individu rép X) / valeur d'alimentation attendue x 100","(Valeur d'alimentation attendue - mm² consommés/jour/individu rép X) x -1 / valeur d'alimentation attendue x 100 --> inhibition = résultat négatif","Constante ache 1 x (Moyenne des masses ^ (Constante ache 2))","Si écart type supérieur au seuil de qualité --> NON sinon OK"," --> inhibition = résultat négatif","Nombre de femelles avec un stade de mue C2 ou D1 et avec des ovocytes","Si std mue fem X = C2 ou D1 alors SI somme des ovocytes > 0 alors somme des ovocytes / taille fem X sinon ""","Nombre de femelles avec un stade de mue C2 ou D1 et avec des ovocytes","Si std emb fem X = 2 ou 3 ou 4 alors SI nombre embryons > 0 alors nombre embryons / (taille fem X - 5) sinon """]:
            ws["A"+row].fill = gray
            ws["A"+row].font = italic
            for letter in header_columns :
                ws[letter+row].fill = gray
                ws[letter+row].font = italic
        if ws["A"+row].value in ["MM² CONSOMMES/JOUR/INDIVIDU - MOYENNE"] :
            ws["A"+row].fill = green
            ws["A"+row].font = bold
            for letter in header_columns :
                ws[letter+row].fill = green
                ws[letter+row].font = bold
        if ws["A"+row].value in ["% INHIBITION (FI) - MOY","% INHIBITION - AChE","% INHIBITION (AChE) - Resultat rendu","% INHIBITION - FERTILITE","% INHIBITION - FECONDITE","% INHIBITION FECONDITE - Resultat rendu"] :
            ws["A"+row].fill = orange
            ws["A"+row].font = bold
            for letter in header_columns :
                ws[letter+row].fill = orange
                ws[letter+row].font = bold
        for letter in header_columns :
            if ws[letter+row].value in ["OK","NON"]:
                ws[letter+row].fill = yellow
        if ws["A"+row].value == "Nombre de femelles analysées" :
            for letter in header_columns :
                if ws[letter+row].value and ws[letter+row].value>15 :
                    ws[letter+row].fill = red
                    ws[letter+row].font = white



        

    wb.save(PATH)
    wb.close()
