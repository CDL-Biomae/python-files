from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter
from termcolor import colored


def add_style_stations(stations_dataframe, PATH):

    wb = load_workbook(PATH)
    ws = wb['Stations']

    nb_rows, nb_columns = stations_dataframe.shape

    medium = Side(border_style='medium', color='FF000000')
    borders = Border(top=medium, left=medium, right=medium, bottom=medium)

    ## COLUMN WIDTH ##
    ws.column_dimensions['A'].width = 13
    ws.column_dimensions['B'].width = 3
    ws.column_dimensions['C'].width = 45
    ws.column_dimensions['D'].width = 13
    ws.column_dimensions['E'].width = 45
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 30
    ws.column_dimensions['I'].width = 27
    ws.column_dimensions['J'].width = 37
    ws.column_dimensions['K'].width = 25
    ws.column_dimensions['L'].width = 37

    ## HEADER STYLE ##
    header_row = '2'
    header_columns = [get_column_letter(col_idx)
                      for col_idx in list(range(1, nb_columns + 1))]
    header_cells = [c+header_row for c in header_columns]

    header_font = Font(size=10, bold=True, name='Arial', color='FFFFFF')
    header_fill = PatternFill(
        fill_type='solid', start_color='808080', end_color='808080')
    header_alignment = Alignment(horizontal='center', vertical='center')

    for cell_str in header_cells:
        cell = ws[cell_str]
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = borders

    ## BODY STYLE ##
    body_rows = [str(r) for r in list(range(3, nb_rows+3))]
    body_columns = header_columns
    body_cells = []
    for row in body_rows:
        for column in body_columns:
            body_cells.append(column+row)

    body_font = Font(size=9, name='Arial')
    body_alignment = Alignment(horizontal='center', vertical='center')

    for cell_str in body_cells:
        cell = ws[cell_str]
        cell.font = body_font
        cell.alignment = body_alignment
        cell.border = borders
    ws.freeze_panes = ws["E3"]
    wb.save(PATH)
    wb.close()

