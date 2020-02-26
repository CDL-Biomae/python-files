from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter
from termcolor import colored
from tools import QueryScript
import math
import env


def add_style_physicochimie(physicochimie_dataframe, filename, folder_PATH):
    PATH = f"{folder_PATH}\\{filename}"
    wb = load_workbook(PATH)
    ws = wb['Physico-chimie_refChimie']
    ws2 = wb['Physico-chimie_refToxicité']

    nb_rows, nb_columns = physicochimie_dataframe.shape

    thin = Side(border_style='thin', color='FF000000')
    medium = Side(border_style='medium', color='FF000000')
    double = Side(border_style='double', color='FF000000')
    no_border = Side(border_style=None)

    font_Arial9 = Font(size=9, name='Arial')
    alignment_center = Alignment(horizontal='center', vertical='center')

    ## COLUMN WIDTH ##
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 45
    ws.column_dimensions['E'].width = 14

    ws2.column_dimensions['A'].width = 3
    ws2.column_dimensions['B'].width = 12
    ws2.column_dimensions['C'].width = 10
    ws2.column_dimensions['D'].width = 45
    ws2.column_dimensions['E'].width = 14

    ## REFORMATAGE STATIONS HEADER ##
    merging_cells = ['B2:B3', 'C2:C3', 'D2:D3', 'E2:E3']
    border_cells = ['B2', 'B3', 'C2', 'C3', 'D2', 'D3', 'E2', 'E3']
    merging_names = ['Campagne', 'Numéro', 'Station de mesure', 'Code Agence']

    header_stations_border = Border(left=medium, right=medium, top=medium, bottom=medium)
    header_stations_font = Font(size=10, bold=True, name='Arial', color='FFFFFF')
    header_stations_fill = PatternFill(fill_type='solid', start_color='808080', end_color='808080')
    header_stations_alignment = Alignment(horizontal='center', vertical='center')

    for worksheet in (ws, ws2):
        for i in range(len(merging_cells)):
            cells = merging_cells[i]
            top_left_cell = worksheet[cells[:2]]
            name = merging_names[i]

            worksheet.merge_cells(cells)
            top_left_cell.value = name
            top_left_cell.font = header_stations_font
            top_left_cell.fill = header_stations_fill
            top_left_cell.alignment = header_stations_alignment

        for cell in border_cells:
            worksheet[cell].border = header_stations_border

    # Nettoyage entre Stations et Values
    cell = ws['F3']
    cell.border = Border(top=no_border, bottom=no_border)
    cell = ws2['F3']
    cell.border = Border(top=no_border, bottom=no_border)

    ## REFORMATAGE VALUE HEADER ##
    # Reformatage temperature
    temperature_cells = 'G2:I2'
    ws.merge_cells(temperature_cells)
    ws2.merge_cells(temperature_cells)

    temp_names = ['Température (°C)', 'minimum', 'moyenne', 'maximum']
    temp_cells = ['G2', 'G3', 'H3', 'I3']
    for worksheet in (ws, ws2):
        for i in range(len(temp_cells)):
            cell = worksheet[temp_cells[i]]
            name = temp_names[i]

            cell.value = name
            cell.font = font_Arial9
            cell.alignment = alignment_center

    temp_border_cells = ['G2', 'H2', 'I2', 'G3', 'H3', 'I3']
    for cell in temp_border_cells:
        if cell[-1] == '2':
            ws[cell].border = Border(top=medium, left=medium, right=medium, bottom=medium)
            ws2[cell].border = Border(top=medium, left=medium, right=medium, bottom=medium)
        else:
            ws[cell].border = Border(top=medium, left=medium, right=medium, bottom=double)
            ws2[cell].border = Border(top=medium, left=medium, right=medium, bottom=double)

    # Reformatage header Conductivité, pH, oxygène
    nb_jours = (nb_columns - 8) // 3
    cond_cells = [f"{get_column_letter(x+10)}2" for x in range(nb_jours)]
    ph_cells = [f"{get_column_letter(x+10)}2" for x in range(nb_jours, nb_jours*2)]
    oxygene_cells = [f"{get_column_letter(x+10)}2" for x in range(nb_jours*2, nb_jours*3)]
    border_cells = cond_cells + ph_cells + oxygene_cells

    ws.merge_cells(f"{cond_cells[0]}:{cond_cells[-1]}")
    ws.merge_cells(f"{ph_cells[0]}:{ph_cells[-1]}")
    ws.merge_cells(f"{oxygene_cells[0]}:{oxygene_cells[-1]}")
    ws2.merge_cells(f"{cond_cells[0]}:{cond_cells[-1]}")
    ws2.merge_cells(f"{ph_cells[0]}:{ph_cells[-1]}")
    ws2.merge_cells(f"{oxygene_cells[0]}:{oxygene_cells[-1]}")

    top_left_cells = [ws[cond_cells[0]], ws[ph_cells[0]], ws[oxygene_cells[0]], ws2[cond_cells[0]], ws2[ph_cells[0]], ws2[oxygene_cells[0]]]
    names = ['Conductivité (µS/cm)', 'pH (unité pH)', 'Oxygène (mg/L)', 'Conductivité (µS/cm)', 'pH (unité pH)', 'Oxygène (mg/L)']

    for i in range(len(top_left_cells)):
        cell = top_left_cells[i]
        name = names[i]

        cell.value = name
        cell.font = font_Arial9
        cell.alignment = alignment_center

    for worksheet in (ws, ws2):
        for cell_str in border_cells:
            cell = worksheet[cell_str]
            cell.border = Border(top=medium, left=medium, right=medium, bottom=medium)

    # Reformatage J#
    J_cells = [f"{get_column_letter(x+10)}3" for x in range(3*nb_jours)]

    for worksheet in (ws, ws2):
        for cell_str in J_cells:
            cell = worksheet[cell_str]
            name = cell.value.split()[-1]

            cell.value = name
            cell.font = font_Arial9
            cell.alignment = alignment_center
            cell.border = Border(top=medium, left=medium, right=medium, bottom=double)


    ## BODY STYLE ##
    # Body Stations
    body_rows = [str(r) for r in range(4, nb_rows + 4)]
    body_stations_columns = ['B', 'C', 'D', 'E']
    body_stations_cells = []
    for row in body_rows:
        for column in body_stations_columns:
            body_stations_cells.append(column+row)

    # Body Values
    body_values_columns = [get_column_letter(x+7) for x in range(3 + 3*nb_jours)]  # 3 temperatures puis 3*nb_jours pour conductivité, ph et oxygene
    body_values_cells = []
    for row in body_rows:
        for column in body_values_columns:
            body_values_cells.append(column+row)

    body_cells = body_stations_cells + body_values_cells

    for worksheet in (ws, ws2):
        for cell_str in body_cells:
            cell = worksheet[cell_str]

            if cell_str[0] == 'H':
                cell.font = Font(size=9, name='Arial', italic=True)
            else:
                cell.font = font_Arial9
            cell.alignment = alignment_center
            cell.border = Border(top=medium, left=medium, right=medium, bottom=medium)


    ## VALUES COMPARAISON TO REFERENCES ##
    # Récupération des références
    output = QueryScript(
        f" SELECT parameter, min, max   FROM {env.DATABASE_TREATED}.r1;"
    ).execute()
    parameters = [x[0] for x in output]
    minimum = [x[1] for x in output]
    maximum = [x[2] for x in output]

    min_temp_chimie = minimum[parameters.index('Température moyenne (chimie)')]
    max_temp_chimie = maximum[parameters.index('Température moyenne (chimie)')]
    min_temp_tox = minimum[parameters.index('Température moyenne (toxicité)')]
    max_temp_tox = maximum[parameters.index('Température moyenne (toxicité)')]

    min_oxygen_chimie = minimum[parameters.index('Oxygène (chimie)')]
    max_oxygen_chimie = maximum[parameters.index('Oxygène (chimie)')]
    min_oxygen_tox = minimum[parameters.index('Oxygène (toxicité)')]
    max_oxygen_tox = maximum[parameters.index('Oxygène (toxicité)')]

    min_cond_chimie = minimum[parameters.index('Conductivité (chimie)')]
    max_cond_chimie = maximum[parameters.index('Conductivité (chimie)')]
    min_cond_tox = minimum[parameters.index('Conductivité (toxicité)')]
    max_cond_tox = maximum[parameters.index('Conductivité (toxicité)')]

    min_pH_chimie = minimum[parameters.index('pH (chimie)')]
    max_pH_chimie = maximum[parameters.index('pH (chimie)')]
    min_pH_tox = minimum[parameters.index('pH (toxicité)')]
    max_pH_tox = maximum[parameters.index('pH (toxicité)')]

    # average temperature
    min_temp_chimie = min_temp_chimie if min_temp_chimie is not None else -math.inf
    max_temp_chimie = max_temp_chimie if max_temp_chimie is not None else math.inf
    min_temp_tox = min_temp_tox if min_temp_tox is not None else -math.inf
    max_temp_tox = max_temp_tox if max_temp_tox is not None else math.inf

    avg_temp_cells = ['H' + x for x in body_rows]

    for cell_str in avg_temp_cells:
        cell = ws[cell_str]
        try:
            avg_temp = float(cell.value)
        except (ValueError, TypeError):
            cell.fill = PatternFill(patternType='solid', start_color='A6A6A6', end_color='A6A6A6')
        else:
            if avg_temp < min_temp_chimie or avg_temp > max_temp_chimie:
                cell.fill = PatternFill(patternType='solid', start_color='B20000', end_color='B20000')

    for cell_str in avg_temp_cells:
        cell = ws2[cell_str]
        try:
            avg_temp = float(cell.value)
        except (ValueError, TypeError):
            cell.fill = PatternFill(patternType='solid', start_color='A6A6A6', end_color='A6A6A6')
        else:
            if avg_temp < min_temp_tox or avg_temp > max_temp_tox:
                cell.fill = PatternFill(patternType='solid', start_color='B20000', end_color='B20000')

    # conductivité
    min_cond_chimie = min_cond_chimie if min_cond_chimie is not None else -math.inf
    max_cond_chimie = max_cond_chimie if max_cond_chimie is not None else math.inf
    min_cond_tox = min_cond_tox if min_cond_tox is not None else -math.inf
    max_cond_tox = max_cond_tox if max_cond_tox is not None else math.inf

    cond_values_columns = [f"{get_column_letter(x + 10)}" for x in range(nb_jours)]
    cond_values_cells = []

    for column in cond_values_columns:
        for row in body_rows:
            cond_values_cells.append(column + row)

    for cell_str in cond_values_cells:
        cell = ws[cell_str]
        try:
            cond = float(cell.value)
        except (ValueError, TypeError):
            cell.fill = PatternFill(patternType='solid', start_color='A6A6A6', end_color='A6A6A6')
        else:
            if cond < min_cond_chimie or cond > max_cond_chimie:
                cell.fill = PatternFill(patternType='solid', start_color='B20000', end_color='B20000')

    for cell_str in cond_values_cells:
        cell = ws2[cell_str]
        try:
            cond = float(cell.value)
        except (ValueError, TypeError):
            cell.fill = PatternFill(patternType='solid', start_color='A6A6A6', end_color='A6A6A6')
        else:
            if cond < min_cond_tox or cond > max_cond_tox:
                cell.fill = PatternFill(patternType='solid', start_color='B20000', end_color='B20000')

    # pH
    min_pH_chimie = min_pH_chimie if min_pH_chimie is not None else -math.inf
    max_pH_chimie = max_pH_chimie if max_pH_chimie is not None else math.inf
    min_pH_tox = min_pH_tox if min_pH_tox is not None else -math.inf
    max_pH_tox = max_pH_tox if max_pH_tox is not None else math.inf

    pH_values_columns = [f"{get_column_letter(x + 10)}" for x in range(nb_jours, nb_jours * 2)]
    pH_values_cells = []

    for column in pH_values_columns:
        for row in body_rows:
            pH_values_cells.append(column + row)

    for cell_str in pH_values_cells:
        cell = ws[cell_str]
        try:
            pH = float(cell.value)
        except (ValueError, TypeError):
            cell.fill = PatternFill(patternType='solid', start_color='A6A6A6', end_color='A6A6A6')
        else:
            if pH < min_pH_chimie or pH > max_pH_chimie:
                cell.fill = PatternFill(patternType='solid', start_color='B20000', end_color='B20000')

    for cell_str in pH_values_cells:
        cell = ws2[cell_str]
        try:
            pH = float(cell.value)
        except (ValueError, TypeError):
            cell.fill = PatternFill(patternType='solid', start_color='A6A6A6', end_color='A6A6A6')
        else:
            if pH < min_pH_tox or pH > max_pH_tox:
                cell.fill = PatternFill(patternType='solid', start_color='B20000', end_color='B20000')

    # oxygene
    min_oxygen_chimie = min_oxygen_chimie if min_oxygen_chimie is not None else -math.inf
    max_oxygen_chimie = max_oxygen_chimie if max_oxygen_chimie is not None else math.inf
    min_oxygen_tox = min_oxygen_tox if min_oxygen_tox is not None else -math.inf
    max_oxygen_tox = max_oxygen_tox if max_oxygen_tox is not None else math.inf

    oxygen_values_columns = [f"{get_column_letter(x + 10)}" for x in range(nb_jours * 2, nb_jours * 3)]
    oxygen_values_cells = []

    for column in oxygen_values_columns:
        for row in body_rows:
            oxygen_values_cells.append(column + row)

    for cell_str in oxygen_values_cells:
        cell = ws[cell_str]
        try:
            oxygen = float(cell.value)
        except (ValueError, TypeError):
            cell.fill = PatternFill(patternType='solid', start_color='A6A6A6', end_color='A6A6A6')
        else:
            if oxygen < min_oxygen_chimie or oxygen > max_oxygen_chimie:
                cell.fill = PatternFill(patternType='solid', start_color='B20000', end_color='B20000')

    for cell_str in oxygen_values_cells:
        cell = ws2[cell_str]
        try:
            oxygen = float(cell.value)
        except (ValueError, TypeError):
            cell.fill = PatternFill(patternType='solid', start_color='A6A6A6', end_color='A6A6A6')
        else:
            if oxygen < min_oxygen_tox or oxygen > max_oxygen_tox:
                cell.fill = PatternFill(patternType='solid', start_color='B20000', end_color='B20000')

    wb.save(PATH)
    wb.close()

    print(colored('[+] La mise en page de l\'onglet \"Physico-chimie\" est terminée', 'green'))
