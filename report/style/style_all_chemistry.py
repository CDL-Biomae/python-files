from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter
import pandas as pd
from tools import QueryScript
from termcolor import colored
import env

def add_style_all_chemistry(all_chemistry_dataframe, PATH, dict_t0):

    wb = load_workbook(PATH)
    ws = wb['Chimie']
    
    nb_rows, nb_columns = all_chemistry_dataframe.shape
    header_row = '4'
    header_columns = [get_column_letter(col_idx) for col_idx in list(range(2, nb_columns + 2))]
    borders = Border(left=Side(border_style='thin', color='FFFFFF'),
                     right=Side(border_style='thin', color='FFFFFF'),
                     top=Side(border_style='thin', color='FFFFFF'),
                     bottom=Side(border_style='thin', color='FFFFFF'))
     
    elements = QueryScript(f" SELECT sandre, parameter, familly FROM {env.DATABASE_TREATED}.r3 WHERE version=  {env.CHOSEN_VERSION()}").execute()
    
    for letter in [get_column_letter(col_idx) for col_idx in range(1, nb_columns+5)]:
        for number in range(1, nb_rows+21):
            ws[letter+str(number)].border = borders
    
    ## Family 
    sandre_list = [element[0] for element in elements]
    parameter_list = [element[1] for element in elements]
    family_list = [element[2] for element in elements]
    for letter in header_columns[5:]:
        index =None
        try :
            if ws[letter + '4'].value and str(ws[letter + '4'].value) in sandre_list:
                index = sandre_list.index(str(ws[letter + '4'].value))
            if index!=None:
                ws[letter + '2'].value = family_list[index]
                ws[letter + '3'].value = sandre_list[index]
                ws[letter + '4'].value = parameter_list[index]
        except ValueError :
            if ws[letter + '4'].value and ws[letter + '4'].value in sandre_list:
                index = sandre_list.index(ws[letter + '4'].value)
            if index!=None:
                ws[letter + '2'].value = family_list[index]
                ws[letter + '3'].value = sandre_list[index]
                ws[letter + '4'].value = parameter_list[index]


    ## Merge unit
    
    current_unit = ws['G2'].value 
    first_letter = 'G'           
    last_letter = 'G'           
    index = 6
    while index <len(header_columns): 
        while index <len(header_columns) and ws[header_columns[index] + '2'].value == current_unit :   
            last_letter = header_columns[index]
            index +=1
        ws.merge_cells(first_letter + '2:'+last_letter+'2')
        if index<len(header_columns):
            first_letter = last_letter = header_columns[index]
        current_unit = ws[first_letter +'2'].value
        index+=1
    
            
    
    ## HEADER STYLE ##
    
    ws['B2'].value = 'Campagne'
    ws['C2'].value = '#'
    ws['D2'].value = 'Station de mesure'
    ws['E2'].value = 'Code agence'
    
    ws.merge_cells('B2:B4')
    ws.merge_cells('C2:C4')
    ws.merge_cells('D2:D4')
    ws.merge_cells('E2:E4')
    
    header_cells = [c+header_row for c in header_columns]
    header_font = Font(size=8, bold=True, name='Arial')
    header_alignment_rotate = Alignment(horizontal='center', vertical='bottom', text_rotation=90)
    header_alignment_no_rotate = Alignment(horizontal='center', vertical='bottom')
    borders = Border(left=Side(border_style='thin', color='FF000000'),
                     right=Side(border_style='thin', color='FF000000'),
                     top=Side(border_style='thin', color='FF000000'),
                     bottom=Side(border_style='thin', color='FF000000'))
    
    ws.column_dimensions['B'].width=3
    ws.column_dimensions['C'].width=3
    ws.column_dimensions['D'].width=30
    ws.column_dimensions['E'].width=8
    for letter in header_columns[:4]:
        for number in range(2,5):
            ws[letter+str(number)].border = borders
            ws[letter+str(number)].font = header_font
    ws['B2'].alignment = header_alignment_rotate
    ws['C2'].alignment = header_alignment_no_rotate
    ws['D2'].alignment = header_alignment_no_rotate
    ws['E2'].alignment = header_alignment_rotate
    
    
    
    header_font = Font(size=8, name='Arial')
    
    for letter in header_columns[4:]:
        if (ws[letter+'5'].value !=None and ws[letter+'5'].value !='') :
            
            ws.column_dimensions[letter].width=6
            ws[letter+'4'].alignment = header_alignment_rotate
            ws[letter+'4'].font = header_font
            ws[letter+'3'].alignment = header_alignment_no_rotate
            ws[letter+'3'].font = header_font
            ws[letter+'2'].alignment = header_alignment_no_rotate
            ws[letter+'2'].font = header_font
        else :
            ws.column_dimensions[letter].width=2
            
    ## ADD T0
    
    t0_mp = []
    for mp in dict_t0:
        if not dict_t0[mp]['code_t0_id'] in t0_mp and dict_t0[mp]['code_t0_id']:
            t0_mp.append(dict_t0[mp]['code_t0_id'])
    if len(t0_mp) > 1 or len(t0_mp) == 0:
        query_tuple_t0 = tuple(t0_mp)
    else:
        query_tuple_t0 = f"({t0_mp[0]})"
    reference_dict = {}
    if len(t0_mp):
        reference_result = QueryScript(f"SELECT reference, id FROM {env.DATABASE_RAW}.Measurepoint WHERE id IN {query_tuple_t0}").execute()
    else:
        reference_result = []
    for reference in reference_result:
        reference_dict.update({reference[1]: reference[0]})
    t0_result = []
    if len(t0_mp):
        t0_result = QueryScript(f"SELECT sandre, prefix, value, Pack.measurepoint_id, Measurepoint.reference FROM {env.DATABASE_RAW}.Analysis JOIN {env.DATABASE_RAW}.Pack ON Pack.id= Analysis.pack_id JOIN {env.DATABASE_RAW}.Measurepoint ON Pack.measurepoint_id=Measurepoint.id WHERE Pack.measurepoint_id IN {query_tuple_t0};").execute()
    dict_t0_result = {}
    for element in t0_result:
        if not element[3] in dict_t0_result:
            dict_t0_result.update({element[3]: {element[0]: element[1] + str(element[2]) if element[1] else str(element[2]), 'reference': element[4]}})
        else:
            dict_t0_result[element[3]][element[0]] = element[1] + str(element[2]) if element[1] else str(element[2])
    
    t0_font = Font(size=6, name='Arial')
    t0_border = Border(left=Side(border_style='thin', color='000000'),
                     right=Side(border_style='thin', color='000000'),
                     top=Side(border_style='thin', color='000000'),
                     bottom=Side(border_style='thin', color='000000'))
    body_alignment = Alignment(horizontal='center', vertical='center')
    
    for index, t0 in enumerate(t0_mp):
        ws['B'+str(nb_rows+5+index)].font = t0_font
        ws['B'+str(nb_rows+5+index)].border = t0_border
        ws['C'+str(nb_rows+5+index)].font = t0_font
        ws['C'+str(nb_rows+5+index)].border = t0_border
        ws['D'+str(nb_rows+5+index)].font = t0_font
        ws['D'+str(nb_rows+5+index)].border = t0_border
        if t0 in dict_t0_result :
            ws['D'+str(nb_rows+5+index)].value = dict_t0_result[t0]['reference']
        else :
            ws['D'+str(nb_rows+5+index)].value = reference_dict[t0]
        ws['E'+str(nb_rows+5+index)].font = t0_font
        ws['E'+str(nb_rows+5+index)].border = t0_border
        for letter in header_columns[5:]:
            sandre = ws[letter +'3'].value
            if t0 in dict_t0_result and  str(sandre) in dict_t0_result[t0]:
                ws[letter+str(nb_rows+5+index)].value = dict_t0_result[t0][str(sandre)]
                ws[letter+str(nb_rows+5+index)].font = t0_font
                ws[letter+str(nb_rows+5+index)].border = t0_border
                ws[letter+str(nb_rows+5+index)].alignment = body_alignment
                
            elif sandre != None :
                ws[letter+str(nb_rows+5+index)].value = 'ND'
                ws[letter+str(nb_rows+5+index)].font = t0_font
                ws[letter+str(nb_rows+5+index)].border = t0_border
                ws[letter+str(nb_rows+5+index)].alignment = body_alignment
        

    ## BODY STYLE ##
    body_rows = [str(r) for r in list(range(5, nb_rows + 5 + len(t0_mp)))]
    body_columns = header_columns[5:]

    body_font = Font(size=6, name='Arial')
    # body_font_white = Font(size=6, name='Arial', color='FFFFFF')
    # body_fill_ok = PatternFill(fill_type='solid', start_color='DBB7FF' if not old_color else '027ee3', end_color='DBB7FF' if not old_color else '027ee3')
    # body_fill_nd = PatternFill(fill_type='solid', start_color='A6A6A6' , end_color='A6A6A6' )
    # body_fill_not_ok = PatternFill(fill_type='solid', start_color='B565F7' if not old_color else '69a64b', end_color='B565F7' if not old_color else '69a64b')
    # body_fill_not_ok_25 = PatternFill(fill_type='solid', start_color='8909FF' if not old_color else 'd1c452', end_color='8909FF' if not old_color else 'd1c452')
    # body_fill_not_ok_50 = PatternFill(fill_type='solid', start_color='6600CC' if not old_color else 'cc7931', end_color='6600CC' if not old_color else 'cc7931')
    # body_fill_not_ok_75 = PatternFill(fill_type='solid', start_color='47008E' if not old_color else 'ab2222', end_color='47008E' if not old_color else 'ab2222')
    # body_alignment = Alignment(horizontal='center', vertical='center')

    for column in header_columns:
        for row in body_rows:
            if column in ['B','C','D','E']:
                ws[column+row].border = borders
                ws[column+row].font = body_font


    #     sandre_checked = ws[column+'3'].value
    #     threshold_7j=''
    #     if sandre_checked!='' and sandre_checked!=None:
    #         try :
    #             index = sandre_list.index(sandre_checked)
    #             threshold_7j = threshold_metal[index][0]
    #             threshold_7j_25 = threshold_metal[index][1]
    #             threshold_7j_50 = threshold_metal[index][2]
    #             threshold_7j_75 = threshold_metal[index][3]
    #         except :
    #             index = sandre_organic.index(sandre_checked)
    #             threshold_7j = threshold_organic[index][0]
    #             threshold_7j_25 = threshold_organic[index][1]
    #             threshold_7j_50 = threshold_organic[index][2]
    #             threshold_7j_75 = threshold_organic[index][3]
    for column in body_columns:
        for row in body_rows:
            cell = ws[column+row]
            value = cell.value
            cell.font = body_font
            if value!=None :
                cell.alignment = body_alignment
                cell.border = borders
                # if (value!="ND" and value!='0.0' and threshold_7j!='') and ((value!='' and value[0]=='<') or float(value)<threshold_7j):
    #                 cell.fill = body_fill_ok
    #             elif (value!="ND" and value!='0.0' and threshold_7j!='' and float(value)>=threshold_7j):
    #                 if threshold_7j_25==None:
    #                     cell.fill = body_fill_not_ok_75
    #                     cell.font = body_font_white
    #                 elif float(value)>=threshold_7j_25 :
    #                     if float(value)>=threshold_7j_50:
    #                         if float(value)>=threshold_7j_75:
    #                             cell.fill = body_fill_not_ok_75
    #                             cell.font = body_font_white
    #                         else :    
    #                             cell.fill = body_fill_not_ok_50 
    #                             cell.font = body_font_white
    #                     else:
    #                         cell.fill = body_fill_not_ok_25
    #                 else :
    #                     cell.fill = body_fill_not_ok
    #             else :
    #                 cell.fill = body_fill_nd
    # for index,mp in enumerate(dict_t0):
    #     try :
    #         if ws[header_columns[-1] + str(index+5)].value:
    #             index_t0_associated = t0_mp.index(dict_t0[ws[header_columns[-1] + str(index+5)].value]['code_t0_id'])
    #             for column in header_columns[5:]:
    #                 t0_ok = True if ws[column + str(5+nb_rows+index_t0_associated)].fill == body_fill_ok else False 
    #                 if not t0_ok and (ws[column + str(5+nb_rows+index_t0_associated)].value!= None and ws[column + str(5+nb_rows+index_t0_associated)].value!= ''):
    #                     ws[column + str(5+index)].fill = body_fill_nd
    #                     ws[column + str(5+index)].font = body_font
    #     except ValueError :
    #         None
    ws.delete_cols(len(header_columns)+1,1)
    
    for letter in header_columns[5:]:
        for number in range(5, nb_rows+21):
                ws[letter + str(number)].value = str(ws[letter + str(number)].value).replace(".", ",") if ws[letter + str(number)].value else ''
    ws.freeze_panes = ws["F4"]
    
    wb.save(PATH)
    wb.close()


