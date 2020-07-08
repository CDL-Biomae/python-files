from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter
import pandas as pd
from calcul import chemistry
from tools import QueryScript
from termcolor import colored
import env

def add_style_chemistry(PATH):

    wb = load_workbook(PATH)
    ws = wb['Data']

    ## COLUMN WIDTH ##
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 6
    ws.column_dimensions['E'].width = 6
    ws.column_dimensions['F'].width = 6
    ws.column_dimensions['G'].width = 30
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 15
    ws.column_dimensions['K'].width = 15
    ws.column_dimensions['L'].width = 15
    ws.column_dimensions['M'].width = 15
    ws.column_dimensions['N'].width = 15
    ws.column_dimensions['O'].width = 15
    ws.column_dimensions['P'].width = 15
    ws.column_dimensions['Q'].width = 15
    ws.column_dimensions['R'].width = 15
    ws.column_dimensions['S'].width = 15
    ws.column_dimensions['T'].width = 15
    ws.column_dimensions['U'].width = 15
    ws.column_dimensions['V'].width = 15
    ws.column_dimensions['W'].width = 15
    ws.column_dimensions['X'].width = 15
    ws.column_dimensions['Y'].width = 15
    ws.column_dimensions['Z'].width = 15
    ws.column_dimensions['AA'].width = 15
    ws.column_dimensions['AB'].width = 15
    ws.column_dimensions['AC'].width = 15
    ws.column_dimensions['AD'].width = 15
    ws.column_dimensions['AE'].width = 15
    ws.column_dimensions['AF'].width = 15
    ws.column_dimensions['AG'].width = 15
    ws.column_dimensions['AH'].width = 15
    ws.column_dimensions['AI'].width = 15
    ws.column_dimensions['AJ'].width = 15
    ws.column_dimensions['AK'].width = 15
    ws.column_dimensions['AL'].width = 15
    ws.column_dimensions['AM'].width = 15
    ws.column_dimensions['AN'].width = 15
    ws.column_dimensions['AO'].width = 15
    ws.column_dimensions['AP'].width = 15
    ws.column_dimensions['AQ'].width = 15
    ws.column_dimensions['AR'].width = 15
    ws.column_dimensions['AS'].width = 15
    ws.column_dimensions['AT'].width = 15

    
    ws.freeze_panes = ws["D2"]
    wb.save(PATH)
    wb.close()


