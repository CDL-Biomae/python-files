from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter
from termcolor import colored

def add_style_physicochimie(physicochimie_dataframe, filename):
    PATH = f"output\\{filename}"
    wb = load_workbook(PATH)
    ws = wb.get_sheet_by_name('Physico-chimie')

    nb_rows, nb_columns = physicochimie_dataframe.shape

