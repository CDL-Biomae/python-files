from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter
import pandas as pd
from calcul import chemistry, elements_crustacean, elements_fish


def add_style_nqe(campagnes_dataframe, filename):
    PATH = f"output\\{filename}"
    wb = load_workbook(PATH)
    ws = wb['NQE Biote']
    
    nb_rows, nb_columns = campagnes_dataframe.shape
    header_row = '4'
    header_columns = [get_column_letter(col_idx) for col_idx in list(range(2, nb_columns + 2))]
    borders = Border(left=Side(border_style='thin', color='FFFFFF'),
                     right=Side(border_style='thin', color='FFFFFF'),
                     top=Side(border_style='thin', color='FFFFFF'),
                     bottom=Side(border_style='thin', color='FFFFFF'))
    
    for letter in  [get_column_letter(col_idx) for col_idx in range(1,nb_columns+2)] :
        for number in range(1,nb_rows+5):
            ws[letter+str(number)].border = borders
    
    ## UNIT 
    # a ordonner
    [unit_crustacean, sandre_crustacean] = chemistry.get_unit(elements_crustacean.keys()) 
    [unit_fish, sandre_fish] = chemistry.get_unit(elements_fish.keys()) 
    print(unit_fish,sandre_fish)
    unit_fish.append('')
    sandre_fish.append('')
    index = 0
    for letter in header_columns[5:]:
        try :
            ws[letter + '2'] = unit_crustacean[index]
            ws[letter + '3'] = sandre_crustacean[index]
            index+=1
        except:
            try : 
                ws[letter + '2'] = unit_fish[index - len(unit_crustacean)-1]
                ws[letter + '3'] = sandre_fish[index - len(unit_crustacean)-1]
                index+=1
            except :
                a=1
                
    
    ## HEADER STYLE ##
    
    ws['B2'].value = 'Campagne'
    ws['C2'].value = '#'
    ws['D2'].value = 'Station de mesure'
    ws['E2'].value = 'Code agence'
    
    ws.merge_cells('B2:B4')
    ws.merge_cells('C2:C4')
    ws.merge_cells('D2:D4')
    ws.merge_cells('E2:E4')
    
    header_cells = [c+header_row for c in header_columns]
    header_font = Font(size=8, bold=True, name='Arial')
    header_alignment_rotate = Alignment(horizontal='center', vertical='bottom', text_rotation=90)
    header_alignment_no_rotate = Alignment(horizontal='center', vertical='bottom')
    borders = Border(left=Side(border_style='thin', color='FF000000'),
                     right=Side(border_style='thin', color='FF000000'),
                     top=Side(border_style='thin', color='FF000000'),
                     bottom=Side(border_style='thin', color='FF000000'))
    
    ws.column_dimensions['B'].width=3
    ws.column_dimensions['C'].width=3
    ws.column_dimensions['D'].width=30
    ws.column_dimensions['E'].width=8
    for letter in header_columns[:4]:
        for number in range(2,5):
            ws[letter+str(number)].border = borders
            ws[letter+str(number)].font = header_font
    ws['B2'].alignment = header_alignment_rotate
    ws['C2'].alignment = header_alignment_no_rotate
    ws['D2'].alignment = header_alignment_no_rotate
    ws['E2'].alignment = header_alignment_rotate
    
    
    
    header_font = Font(size=8, name='Arial')
    
    for letter in header_columns[4:]:
        if (ws[letter+'5'].value !=None and ws[letter+'5'].value !='') :
            
            ws.column_dimensions[letter].width=5
            ws[letter+'4'].alignment = header_alignment_rotate
            ws[letter+'4'].font = header_font
            ws[letter+'3'].alignment = header_alignment_no_rotate
            ws[letter+'3'].font = header_font
            ws[letter+'2'].alignment = header_alignment_no_rotate
            ws[letter+'2'].font = header_font
        else :
            ws.column_dimensions[letter].width=2
            
    
    
    # for letter in header_columns:
    #     for i in range(2,6):
    #         ws[letter+str(i)].font = header_font
    #         ws[letter+str(i)].border = borders
    #         ws[letter+str(i)].alignment = header_alignment_measure
    
    # for cell_str in header_cells:
    #     cell = ws[cell_str]
    #     if cell.value=='Station de mesure':
    #         cell.alignment = header_alignment_measure
    #         cell.border = borders
    #     elif cell.value!=None :
    #         cell.border = borders
    #         cell.alignment = header_alignment


    ## BODY STYLE ##
    body_rows = [str(r) for r in list(range(5, nb_rows+5))]
    body_columns = header_columns
    body_cells = []
    for row in body_rows:
        for column in body_columns:
            body_cells.append(column+row)

    body_font = Font(size=6, name='Arial')
    body_fill_ND = PatternFill(fill_type='solid', start_color='606060', end_color='606060')
    body_fill_not_ND = PatternFill(fill_type='solid', start_color='318CE7', end_color='318CE7')
    body_alignment = Alignment(horizontal='center', vertical='center')

    for cell_str in body_cells:
        cell = ws[cell_str]
        value = cell.value
        cell.font = body_font
        if value!=None:
            cell.alignment = body_alignment
            cell.border = borders
            if value=="ND":
                cell.fill = body_fill_ND
            elif cell.column>"F":
                cell.fill = body_fill_not_ND
                

    wb.save(PATH)
    wb.close()

    print('[+] La mise en page de l\'onglet \"NQE\" est terminée')

