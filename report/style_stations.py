from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter
from termcolor import colored


def add_style_stations(stations_dataframe, filename):
    PATH = f"output\\{filename}"
    wb = load_workbook(PATH)
    ws = wb.get_sheet_by_name('Stations')

    nb_rows, nb_columns = stations_dataframe.shape

    borders = Border(left=Side(border_style='thin', color='FF000000'),
                     right=Side(border_style='thin', color='FF000000'),
                     top=Side(border_style='thin', color='FF000000'),
                     bottom=Side(border_style='thin', color='FF000000'))

    ## HEADER STYLE ##
    header_row = '2'
    header_columns = [get_column_letter(col_idx) for col_idx in list(range(2, nb_columns + 2))]
    header_cells = [c+header_row for c in header_columns]

    header_font = Font(size=10, bold=True, name='Arial', color='FFFFFF')
    header_fill = PatternFill(fill_type='solid', start_color='808080', end_color='808080')
    header_alignment = Alignment(horizontal='center', vertical='center')

    for cell_str in header_cells:
        cell = ws[cell_str]
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = borders


    ## BODY STYLE ##
    body_rows = [str(r) for r in list(range(3, nb_rows+3))]
    body_columns = header_columns
    body_cells = []
    for row in body_rows:
        for column in body_columns:
            body_cells.append(column+row)

    body_font = Font(size=9, name='Arial')
    body_alignment = Alignment(horizontal='center', vertical='center')

    for cell_str in body_cells:
        cell = ws[cell_str]
        cell.font = body_font
        cell.alignment = body_alignment
        cell.border = borders

    wb.save(PATH)
    wb.close()

    print(colored('[+] La mise en page de l\'onglet \"Stations\" est terminée', 'green'))

